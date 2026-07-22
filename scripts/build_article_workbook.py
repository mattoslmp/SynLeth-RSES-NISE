#!/usr/bin/env python3
"""Build the complete organized article, supplementary, QC and evidence workbook."""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
EXCEL_MAX_ROWS = 1_048_576


def resolve_path(value: str) -> Path:
  path = Path(value)
  return path if path.is_absolute() else ROOT / path


def safe_sheet_name(name: str, used: set[str]) -> str:
  base = name[:31]
  candidate = base
  suffix = 1
  while candidate in used:
    trailer = f"_{suffix}"
    candidate = base[: 31 - len(trailer)] + trailer
    suffix += 1
  used.add(candidate)
  return candidate


def format_workbook(path: Path) -> None:
  workbook = load_workbook(path)
  header_fill = PatternFill("solid", fgColor="17365D")
  header_font = Font(color="FFFFFF", bold=True, size=10)
  title_fill = PatternFill("solid", fgColor="DDEBF7")
  for worksheet in workbook.worksheets:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
      cell.fill = header_fill
      cell.font = header_font
      cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 36
    for column_index, column_cells in enumerate(worksheet.columns, start=1):
      maximum = 0
      for cell in list(column_cells)[:300]:
        value = "" if cell.value is None else str(cell.value)
        maximum = max(maximum, max((len(line) for line in value.splitlines()), default=0))
        if cell.row > 1:
          cell.alignment = Alignment(vertical="top", wrap_text=True)
      worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(maximum + 2, 11), 55)
  if "Contents" in workbook.sheetnames:
    contents = workbook["Contents"]
    for cell in contents[1]:
      cell.fill = title_fill
      cell.font = Font(bold=True, color="17365D")
  workbook.save(path)


def table_groups(article_root: Path) -> list[tuple[str, list[Path]]]:
  groups = [
    ("Main", sorted((article_root / "tables/main").glob("*.tsv"))),
    ("Supplementary", sorted((article_root / "tables/supplementary").glob("*.tsv"))),
    ("QC", sorted((article_root / "tables/qc").glob("*.tsv"))),
    ("Score components", sorted((article_root / "tables/score_components").glob("*.tsv"))),
    ("Robustness", sorted((article_root / "tables/robustness").glob("*.tsv"))),
    ("Figure data", sorted((article_root / "tables/figure_data").glob("*.tsv"))),
    ("Supporting evidence", sorted((article_root / "tables/supporting_evidence").rglob("*.tsv"))),
    ("Manifest", sorted((article_root / "manifests").glob("*.tsv"))),
  ]
  return [(category, paths) for category, paths in groups if paths]


def write_frame_chunks(
  writer: pd.ExcelWriter,
  frame: pd.DataFrame,
  path: Path,
  category: str,
  used: set[str],
  contents_rows: list[dict[str, object]],
) -> None:
  maximum_data_rows = EXCEL_MAX_ROWS - 1
  if frame.empty:
    chunks = [(1, frame)]
  else:
    chunks = [
      (index + 1, frame.iloc[start:start + maximum_data_rows])
      for index, start in enumerate(range(0, len(frame), maximum_data_rows))
    ]
  for part, chunk in chunks:
    suffix = f"_p{part}" if len(chunks) > 1 else ""
    sheet = safe_sheet_name(path.stem + suffix, used)
    chunk.to_excel(writer, sheet_name=sheet, index=False)
    contents_rows.append({
      "category": category,
      "sheet": sheet,
      "source_file": str(path),
      "part": part,
      "rows_in_sheet": len(chunk),
      "total_source_rows": len(frame),
      "columns": len(frame.columns),
      "source_status": "available" if path.exists() else "missing",
    })


def main() -> None:
  parser = argparse.ArgumentParser()
  parser.add_argument("--article-root", default="article_outputs")
  parser.add_argument(
    "--output",
    default="article_outputs/workbooks/RSES_Onco_Article_Tables_and_Evidence.xlsx",
  )
  args = parser.parse_args()
  article_root = resolve_path(args.article_root)
  output = resolve_path(args.output)
  output.parent.mkdir(parents=True, exist_ok=True)
  groups = table_groups(article_root)
  main_tables = dict(groups).get("Main", [])
  supplementary_tables = dict(groups).get("Supplementary", [])
  if not main_tables or not supplementary_tables:
    raise RuntimeError("Article tables are absent. Run scripts/export_article_tables.py first.")

  contents_rows: list[dict[str, object]] = []
  used: set[str] = set()
  with pd.ExcelWriter(output, engine="openpyxl") as writer:
    for category, paths in groups:
      for path in paths:
        frame = pd.read_csv(path, sep="\t", low_memory=False)
        write_frame_chunks(writer, frame, path, category, used, contents_rows)
    contents = pd.DataFrame(contents_rows)
    contents.to_excel(writer, sheet_name="Contents", index=False)
  format_workbook(output)
  if not output.exists() or output.stat().st_size < 1000:
    raise RuntimeError(f"Workbook was not produced or is too small: {output}")
  print(f"Workbook sheets: {len(contents_rows) + 1}")
  print(f"Wrote {output}")


if __name__ == "__main__":
  main()
